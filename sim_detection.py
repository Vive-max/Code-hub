import subprocess
import os
import time
from xlwt import Workbook
from openpyxl import Workbook
import sys
from openpyxl import load_workbook
id = subprocess.getstatusoutput("adb devices")
print(id)

sid=id
k = sid[1].split("\n")
v=k[1]
oop = v.split("\tdevice")
print(oop)
xy = oop[0]
print(xy)
s = str(input("######## PLZ scan the devices id--########\n",))
p = (s[3:12])
key=s[18:26]
print("kreat key",key)

vk = subprocess.getstatusoutput("adb -s d6217f12  shell echo ssid=DCAM-{config} /etc/misc/wifi/hostapd-open.conf".format(config=p))
time.sleep(1)
s = subprocess.getstatusoutput("adb -s {deviceid}  shell systemctl stop dcam_wifi".format(deviceid=xy))
print(s)

time.sleep(1)
h = subprocess.getstatusoutput("adb -s {device}  shell wifi_ap.sh".format(device=xy))
print(h)

os.system("nmcli d wifi rescan")
time.sleep(4)
os.system("nmcli d wifi connect DCAM-{enterid}".format(enterid=p))
os.system('adb tcpip 5555')
print("@@@@#######--plz disconnect the usb cable--@@@@@####")

time.sleep(4)

os.system('adb connect 192.168.43.1:5555')

time.sleep(10)

n=0
while n<3:

        lumia = subprocess.getstatusoutput("adb -s 192.168.43.1:5555 shell lsusb")
        print(lumia)
        n = n+1

wq= subprocess.getstatusoutput("adb -s 192.168.43.1:5555 shell lsusb")
print(wq)
time.sleep(1)
os.system("adb -s 192.168.43.1:5555 shell lsusb")
time.sleep(1)
yd = subprocess.getstatusoutput(" adb -s 192.168.43.1:5555 shell lte_gps_sample_app 'at!iccid?' ")
print(yd)
sub=(0, 'Err,Sierra modem not detected and Hence usb ports are also not detected')
if sub==yd:
    g=0
    while g<70:

        wq = subprocess.getstatusoutput("adb -s 192.168.43.1:5555 shell lsusb")
        print("plz wait....")
        g=g+1
else:
    pass
os.system("adb -s 192.168.43.1:5555 shell lsusb")
yd = list(subprocess.getstatusoutput(" adb -s 192.168.43.1:5555 shell lte_gps_sample_app 'at!iccid?' "))
print(yd)
sim_id = (str(yd[1]).split("modem_response: \n\nat!iccid?\n\n\n!ICCID: ")[1])[:20]
print("sim_id",sim_id)
can = str(input("#########---plz scan the CAN id--##########\n",))
canid = can[3:12]
print("can_id",canid)

LS = str(input("#######--plz scan the LUMIA id---########\n",))
LumiaSI = LS[0:9]
print("lumia serial number",LumiaSI)

fru = subprocess.getstatusoutput('adb -s 192.168.43.1:5555 shell read_kraitinfo_emmc')
print(fru)
alldata = fru[1].split("\n")

Inboard = alldata[1]

boardno = Inboard.split("Mainboard serial_num:")

b = boardno[1]
print("MainBoard serial number",b)

productno = alldata[-3]
pNO= productno.split("Product serial_num:")
p=pNO[1]
print("Boxboard serial_num",p)

IMEIno = alldata[-1]
imeiNO = IMEIno.split("IMEI serial_num:")
emi = imeiNO[1]
print("lumia IMEI number",emi)

o = alldata[2]
out = o.split("OutCam version:")
ow = out[1]
print("outcam_version",ow)

i = alldata[3]
inw = i.split("InCam version:")
inwoard = inw[1]
print("incam version",inwoard)

ir = alldata[4]
irl = ir.split("IRLED board serial_num:")
irled = irl[1]
print("irled serial number",irled)

m = s[34:41]
print("kreat_model_id",m)

SKU=s[47:57]
print("kreat_sku",SKU)

print("@@@@@##########@@@@@-DONE-@@@@@##########@@@@@-DONE-@@@@@@################@@@@@@-DONE-@@@@@##############@@@@@-DONE-@@@@@############@@@@@-DONE-@@@@@@##################@@@@@-DONE-@@@@@@###########@@@@@---THANK YOU--@@@@@@##########################@@@@@-DONE-@@@@@@@##############################@@@@@-DONE-@@@@@################################@@@@@-DONE-@@@@@@@######################@@@@@-DONE-@@@@#################################@@@@@-DONE-@@@@@#########################@@@@@-DONE-@@@@@############################@@@-DONE-@@@@@#############################@@@@@-DONE-@@@@@############################@@@@-DONE-@@@@##################################@@@@-DONE-@@@@##############@@@@-DONE-@@@#################@@@@---THANK YOU---@@@#############@@@@@@@@@")

wb = load_workbook(filename = 'abc.xlsx')
book = Workbook()
sheet = book.active

sheet['A1'] = 'Sim_id'
sheet['B1'] = 'Main_board'
sheet['C1'] = 'BOARD_NUMBER'
sheet['D1'] = 'IMEI_NO'
sheet['E1'] = 'LUMIA_ID'
sheet['F1'] = 'OUT_WARD_SI'
sheet['G1'] = 'IN_WOARD'
sheet['H1'] = 'IRLED_NUMBER'
sheet['I1'] = 'CAN_ID'
sheet['J1'] = 'KREAT_KEY'
sheet['K1'] = 'KREAT_MODEL'
sheet['L1'] = 'KREAT_SKU'

sheet.cell(row=4, column=1).value = sim_id

sheet.cell(row=4, column=2).value = b

sheet.cell(row=4, column=3).value = p

sheet.cell(row=4, column=4).value = emi

sheet.cell(row=4, column=5).value = LumiaSI

sheet.cell(row=4, column=6).value = ow

sheet.cell(row=4, column=7).value = inwoard

sheet.cell(row=4, column=8).value = irled

sheet.cell(row=4, column=9).value = canid

sheet.cell(row=4, column=10).value = key

sheet.cell(row=4, column=11).value = m

sheet.cell(row=4, column=12).value = SKU

book.save('abc.xlsx')

os.system("adb disconnect 192.168.43.1:5555")

os.system("sudo adb kill-server")
os.system("sudo adb start-server")
